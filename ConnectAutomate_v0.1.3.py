import tkinter as tk
from tkinter import filedialog, messagebox,simpledialog
from tkinter import ttk
from tkinter.ttk import Combobox
import zipfile
import xml.etree.ElementTree as ET
import json
import os
from docx import Document
import openpyxl
import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

class Field:
    def __init__(self):
        self.id = 0

        self.label = ""
        self.value = ""
        self.type  = "" # Word or Excel
        self.path  = ""

        self.sheetName = ""

        self.tkLabel = None
        self.tkInput = None
        self.tkOption = None
        self.tkDelete = None

        self.isDateToday = False
        self.isSameValue = -1
        self.isSameExcelAs = None
        self.isList = []

        self.pathToSave = ""
        self.wordNameToSave = ""
        self.subfolderName = ""
        self.genericPrefix = -1


class BookmarkForm:
    def __init__(self):
        self.root = tk.Tk()

        self.fields = []

        self.form_frame = None
        self.entries = {}
        self.settings_window = {}
        self.DateDuJour_buttons = {}
        self.linked_vars = {}
        self.master_vars = {}
        self.date_actuelle_signets = set()
        self.list_value = {}

        self.load_word_file_btn = None
        self.add_field_btn = None
        self.add_mail_btn = None
        self.save_btn = None
        self.modify_btn = None
        self.see_mail_btn = None
        self.start_fill_btn = None
        self.form_canvas = None


        self.close_btn = {}
        self.labels = {}

        self.askToSave = False

        self.root.title("ConnectAutomate")
        self.root.geometry("1000x800")  # Largeur x Hauteur

        self.tmp_value = ""
        self.currId = 0
        self.currSheetName = ""

        self.mdp = ""

        self.mail_info = {}  
        self.save_form_infos = []
        self.save_form_index = 0

        # Style commun pour tous les boutons
        self.button_style = {'fg': "darkred", 'bg': "white", 'font': ("Helvetica", 12, "bold")}  

    def load_saved_form(self):
        save_dir = os.path.join(os.path.dirname(__file__), "formulaires")
        filepath = filedialog.askopenfilename(initialdir=save_dir, title="Choisir un formulaire sauvegardé", filetypes=[("JSON files", "*.json")])
        if filepath:
            with open(filepath, 'r') as f:
                form_data = json.load(f)
            
            form_data_fields = form_data["fields"]
            if "mail_info" in form_data:
                    self.mail_info = form_data["mail_info"]

            for data in form_data_fields:
                self.create_field(data["label"], data["path"], data["type"], data["value"], data["id"], data["isDateToday"], data["isSameValue"], data["isSameExcelAs"], data["sheetName"], data["isList"], data["pathToSave"], data["wordNameToSave"], data["subfolderName"], data["genericPrefix"])

            self.mdp = form_data["passWordForm"]

            self.create_new_form()
            self.create_form_2("load")
   
    def open_save_form_data_change_path(self, entry):
        file_path = filedialog.askdirectory()

        if file_path:
            entry.config(state=tk.NORMAL)
            entry.delete(0, tk.END)
            entry.insert(0, file_path)
            entry.config(state='readonly', readonlybackground='light grey')

    def open_save_form_data_change_subfolder_state(self, subfolder_entry, state):
        if state == "Yes":
            subfolder_entry.config(state=tk.NORMAL)
        else:
            subfolder_entry.config(state='readonly', readonlybackground='light grey')

    def open_save_form_data(self):
        # Créer une nouvelle fenêtre ou un cadre pour les paramètres
        save_form_data = tk.Toplevel(self.root)
        save_form_data.title("Save form")
        save_form_data.geometry("1200x700")  # Définir la taille de la fenêtre

        # Create a Canvas widget
        canvas = tk.Canvas(save_form_data)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Add a Scrollbar to the Toplevel window
        scrollbar = tk.Scrollbar(save_form_data, orient=tk.VERTICAL, command=canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Configure the Canvas scrolling behavior
        canvas.configure(yscrollcommand=scrollbar.set)

        # Create a frame within the Canvas
        frame = tk.Frame(canvas)
        canvas.create_window((0, 0), window=frame, anchor="nw")

        path_form_label = tk.Text(frame, height=1, width=50, borderwidth=0, highlightthickness=0)
        path_form_label.insert(tk.END, "Where to save the form")
        path_form_label.config(state=tk.DISABLED)
        path_form_label.grid(row=0, column=0)

        path_form_entry = tk.Entry(frame, width=50)
        path_form_entry.config(state='readonly', readonlybackground='light grey')
        path_form_entry.grid(row=0, column=1)

        path_form_os = tk.Button(frame, text="...", command=lambda: self.open_save_form_data_change_path(path_form_entry))
        path_form_os.grid(row=0, column=2)



        name_form_label = tk.Text(frame, height=1, width=50, borderwidth=0, highlightthickness=0)
        name_form_label.insert(tk.END, "Name of the form")  
        name_form_label.config(state=tk.DISABLED)
        name_form_label.grid(row=1, column=0)

        name_form_entry = tk.Entry(frame, width=50)
        name_form_entry.grid(row=1, column=1)



        password_form_label = tk.Text(frame, height=1, width=50, borderwidth=0, highlightthickness=0)
        password_form_label.insert(tk.END, "Password of the form")
        password_form_label.config(state=tk.DISABLED)
        password_form_label.grid(row=2, column=0)

        password_form_entry = tk.Entry(frame, width=50)
        password_form_entry.grid(row=2, column=1)


        words_path = []
        for field in self.fields:
            if field.type == "Word":
                if field.path not in words_path:
                    words_path.append(field.path)

        current_row = 3

        self.save_form_infos = []
        self.save_form_index = 0
        for path in words_path:
            self.save_form_infos.append({})
    
            self.save_form_infos[self.save_form_index]['originalPath'] = path

            ####  FOLDER WORD SAVE  ####
            self.save_form_infos[self.save_form_index]['wordPathLabel'] = tk.Text(frame, height=1, width=50, borderwidth=0, highlightthickness=0)
            self.save_form_infos[self.save_form_index]['wordPathLabel'].insert(tk.END, f"Where to save {os.path.basename(path)}")
            self.save_form_infos[self.save_form_index]['wordPathLabel'].config(state=tk.DISABLED)
            self.save_form_infos[self.save_form_index]['wordPathLabel'].grid(row=current_row, column=0)

            self.save_form_infos[self.save_form_index]['wordPathEntry'] = tk.Entry(frame, width=50)
            self.save_form_infos[self.save_form_index]['wordPathEntry'].config(state='readonly', readonlybackground='light grey')
            self.save_form_infos[self.save_form_index]['wordPathEntry'].grid(row=current_row, column=1)

            self.save_form_infos[self.save_form_index]['wordPathButton'] = tk.Button(frame, text="...", command=lambda index=self.save_form_index: self.open_save_form_data_change_path(self.save_form_infos[index]['wordPathEntry']))
            self.save_form_infos[self.save_form_index]['wordPathButton'].grid(row=current_row, column=2)


            ####  SUBFOLDER NAME  ####
            self.save_form_infos[self.save_form_index]['subfolderNameLabel'] = tk.Text(frame, height=1, width=20, borderwidth=0, highlightthickness=0)
            self.save_form_infos[self.save_form_index]['subfolderNameLabel'].insert(tk.END, "Subfolder name")
            self.save_form_infos[self.save_form_index]['subfolderNameLabel'].config(state=tk.DISABLED)
            
            self.save_form_infos[self.save_form_index]['subfolderNameEntry'] = tk.Entry(frame, width=30)
            self.save_form_infos[self.save_form_index]['subfolderNameEntry'].config(state='readonly', readonlybackground='light grey')

            self.save_form_infos[self.save_form_index]['subfolderNameRadioSelect'] = tk.StringVar(value="No")
            self.save_form_infos[self.save_form_index]['subfolderNameRadio1'] = tk.Radiobutton(frame, text="Yes", variable=self.save_form_infos[self.save_form_index]['subfolderNameRadioSelect'], value="Yes", command=lambda index=self.save_form_index: self.open_save_form_data_change_subfolder_state(self.save_form_infos[index]['subfolderNameEntry'], "Yes"))
            self.save_form_infos[self.save_form_index]['subfolderNameRadio1'].grid(row=current_row, column=3)

            self.save_form_infos[self.save_form_index]['subfolderNameRadio2'] = tk.Radiobutton(frame, text="No", variable=self.save_form_infos[self.save_form_index]['subfolderNameRadioSelect'], value="No", command=lambda index=self.save_form_index: self.open_save_form_data_change_subfolder_state(self.save_form_infos[index]['subfolderNameEntry'], "No"))
            self.save_form_infos[self.save_form_index]['subfolderNameRadio2'].grid(row=current_row, column=4)

            self.save_form_infos[self.save_form_index]['subfolderNameLabel'].grid(row=current_row, column=5)
            self.save_form_infos[self.save_form_index]['subfolderNameEntry'].grid(row=current_row, column=6)



            ####  WORD NAME  ####
            self.save_form_infos[self.save_form_index]['wordNameLabel'] = tk.Text(frame, height=1, width=50, borderwidth=0, highlightthickness=0)
            self.save_form_infos[self.save_form_index]['wordNameLabel'].insert(tk.END, "Name of the word")
            self.save_form_infos[self.save_form_index]['wordNameLabel'].config(state=tk.DISABLED)
            self.save_form_infos[self.save_form_index]['wordNameLabel'].grid(row=current_row+1, column=0)

            self.save_form_infos[self.save_form_index]['wordNameEntry'] = tk.Entry(frame, width=50)
            self.save_form_infos[self.save_form_index]['wordNameEntry'].grid(row=current_row+1, column=1)



            ####  GENERIC ADD TO WORD NAME  ####
            self.save_form_infos[self.save_form_index]['genericNameLabel'] = tk.Text(frame, height=1, width=50, borderwidth=0, highlightthickness=0)
            self.save_form_infos[self.save_form_index]['genericNameLabel'].insert(tk.END, "Generic add to the word name")
            self.save_form_infos[self.save_form_index]['genericNameLabel'].config(state=tk.DISABLED)
            self.save_form_infos[self.save_form_index]['genericNameLabel'].grid(row=current_row+2, column=0)

            self.save_form_infos[self.save_form_index]['genericComboboxString'] = tk.StringVar()
            self.save_form_infos[self.save_form_index]['genericCombobox'] = ttk.Combobox(frame, textvariable=self.save_form_infos[self.save_form_index]['genericComboboxString'], width=50, state="readonly")

            self.save_form_infos[self.save_form_index]['genericSelectChoices'] = [""]
            self.save_form_infos[self.save_form_index]['genericIds'] = [-1]
            for field in self.fields:
                self.save_form_infos[self.save_form_index]['genericSelectChoices'].append(field.label)
                self.save_form_infos[self.save_form_index]['genericIds'].append(field.id)
            
            self.save_form_infos[self.save_form_index]['genericCombobox']['values'] = self.save_form_infos[self.save_form_index]['genericSelectChoices']
            self.save_form_infos[self.save_form_index]['genericCombobox'].grid(row=current_row+2, column=1)


            ####  END  ####
            current_row += 3
            self.save_form_index += 1

        
        save_button = tk.Button(frame, text="Save data", command=lambda: 
                                self.save_form_data(
                                    path_form_entry,
                                    name_form_entry,
                                    password_form_entry
                                ), **self.button_style)
        save_button.grid(row=current_row, column=1)

    def save_form_add_infos_fields(self):
        for field in self.fields:
            for word in self.save_form_infos:
                if word['originalPath'] == field.path:
                    fieldOfGeneric = self.find_field_with_id(word['genericIds'][word['genericCombobox'].current()])
                    if fieldOfGeneric != None:
                        field.genericPrefix = fieldOfGeneric.id
                    else:
                        field.genericPrefix = -1

                    if word['wordPathEntry'].get() == "" or word['wordNameEntry'].get() == "":
                        return os.path.basename(word['originalPath'])

                    field.pathToSave = word['wordPathEntry'].get() #f"{word['wordPathEntry'].get()}/{word['wordNameEntry'].get()}_{fieldOfGenericStr}"
                    field.wordNameToSave = f"{word['wordNameEntry'].get()}" #_{fieldOfGenericStr}"
                    if word['subfolderNameEntry']['state'] != "readonly":
                        field.subfolderName = f"{word['subfolderNameEntry'].get()}" #_{fieldOfGenericStr}"
                        #field.subfolderName = f"{word['wordPathEntry'].get()}/{word['subfolderNameEntry'].get()}_{fieldOfGenericStr}"

        return ""

    def save_form_data(self, path, name, password):
        path_str = path.get()
        name_str = name.get()
        pass_str = password.get()

        if path_str == "":
            messagebox.showinfo("Error", "Error in the path of the form")
            return
        if name_str == "":
            messagebox.showinfo("Error", "Error in the name of the form")
            return
    
        error_res = self.save_form_add_infos_fields()
        if error_res != "":
            messagebox.showinfo("Error", f"Error in {error_res}")
            return

        # Chemin complet du fichier de sauvegarde
        filepath = os.path.join(path_str, f"{name_str}.json")

        tmp = {
            "fields": [],
            "mail_info": self.mail_info,
            "passWordForm": pass_str
        }
        for value in self.fields:
            tmp["fields"].append({
                "id": value.id,
                "label": value.label,
                "value": value.tkInput.get(),
                "type": value.type,
                "path": value.path,
                "isDateToday": value.isDateToday,
                "isSameValue": value.isSameValue,
                "isSameExcelAs": value.isSameExcelAs,
                "sheetName": value.sheetName,
                "isList": value.isList,
                "pathToSave": value.pathToSave,
                "wordNameToSave": value.wordNameToSave,
                "subfolderName": value.subfolderName,
                "genericPrefix": value.genericPrefix
            })

        # Écrire les données dans un fichier JSON
        with open(filepath, 'w') as f:
            json.dump(tmp, f, indent=4)

        messagebox.showinfo("Sauvegarde réussie", "Le formulaire a été sauvegardé.")
        self.askToSave = False
        self.reset_to_menu()

    def create_form_2(self, state):
        self.askToSave = True

        # Pour supprimer/masquer les boutons
        if self.load_word_file_btn != None:
            self.load_word_file_btn.grid_forget()
        if self.add_field_btn != None:
            self.add_field_btn.grid_forget()
        if self.add_mail_btn != None:
            self.add_mail_btn.grid_forget()
        if self.save_btn != None:
            self.save_btn.grid_forget()
        if self.modify_btn != None:
            self.modify_btn.grid_forget()
        if self.see_mail_btn != None:
            self.see_mail_btn.grid_forget()
        if self.start_fill_btn != None:
            self.start_fill_btn.grid_forget()

        # Créer un nouveau cadre pour les éléments de formulaire
        if self.form_frame is None:
            self.form_canvas = tk.Canvas(self.root)
            self.form_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            scrollbar = tk.Scrollbar(self.root, orient=tk.VERTICAL, command=self.form_canvas.yview)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # Configure the Canvas scrolling behavior
            self.form_canvas.configure(yscrollcommand=scrollbar.set)

            self.form_frame = tk.Frame(self.form_canvas)
            #self.form_frame.pack(fill='x', pady=20)
            self.form_canvas.create_window((0, 0), window=self.form_frame, anchor="nw")

            def on_configure(event):
                self.form_canvas.configure(scrollregion=self.form_canvas.bbox("all"))
            self.form_canvas.bind("<Configure>", on_configure)

        for index, value in enumerate(self.fields):
            if value.tkLabel == None:
                # Label
                text_widget = tk.Text(self.form_frame, height=1, width=50, borderwidth=0, highlightthickness=0)
                text_widget.grid(row=index, column=0)
                prefix = "Bookmark :" if value.type == "Word" else "Column :"
                text_widget.insert(tk.END, f"{prefix} {value.label}:")
                text_widget.tag_add("prefix", "1.0", f"1.{len(prefix)}")
                text_widget.tag_configure("prefix", foreground="#00008B" if value.type == "Word" else "#006400", font=('Arial', -12, 'italic'))
                text_widget.config(state=tk.DISABLED)
                value.tkLabel = text_widget

                # Text Input
                entry = tk.Entry(self.form_frame)
                entry.grid(row=index , column=1)
                entry.insert(0, value.value)
                value.tkInput = entry

                # Option Button
                settings_btn = tk.Button(self.form_frame, text="⋮", command=lambda id=value.id: self.open_settings(id))
                settings_btn.grid(row=index, column=2)
                value.tkOption = settings_btn
                if state == "load":
                    value.tkOption.grid_forget()

                # Delete Button
                close_btn = tk.Button(self.form_frame, text="x", command=lambda id=value.id: self.close_button(id))
                close_btn.grid(row=index, column=3)
                value.tkDelete = close_btn
                if state == "load":
                    value.tkDelete.grid_forget()

                if value.isSameValue != -1:
                    self.link_variable(value.id, value.isSameValue)
                
                if value.isDateToday:
                    self.set_Date_actuelle(value.id)

                if value.isSameExcelAs != None:
                    self.set_as_student(value.id, value.isSameExcelAs)

                if value.isList != [] :
                    self.set_list_of_value(value.id, value.isList)
                    
                    

        if state == "modify":
            # Bouton pour charger un fichier Word
            self.load_word_file_btn = tk.Button(self.form_frame, text="Automatic filling from a Word, Excel file", command=self.load_file, fg="darkred", bg="white", font=("Helvetica", 12, "bold"))
            self.load_word_file_btn.grid(row=len(self.fields), column=1, padx=10, pady=20)

            # Bouton pour ajouter un nouveau champ
            self.add_field_btn = tk.Button(self.form_frame, text="+ Add a new field", command=self.add_field, fg="darkred", bg="white", font=("Helvetica", 12, "bold"))  # Assurez-vous que la commande est correcte ici
            self.add_field_btn.grid(row=len(self.fields) + 1, column=1, padx=10, pady=20)

            # Placer le bouton de sauvegarde sous ces boutons
            button_text = "Modify mail" if self.mail_info else "Add a mail"
            self.add_mail_btn = tk.Button(self.form_frame, text=button_text, command=self.add_mail, fg="darkred", bg="white", font=("Helvetica", 12, "bold"))
            self.add_mail_btn.grid(row=len(self.fields) + 2, column=1, padx=10, pady=20)  # Utilisez columnspan pour centrer le bouton si vous le souhaitez

            # Placer le bouton de sauvegarde sous ces boutons
            self.save_btn = tk.Button(self.form_frame, text="Save the form", command=self.open_save_form_data, fg="darkred", bg="white", font=("Helvetica", 12, "bold"))
            self.save_btn.grid(row=len(self.fields) + 3, column=1, padx=10, pady=20)  # Utilisez columnspan pour centrer le bouton si vous le souhaitez
        elif state == "load":
            # Bouton pour modifier le form
            self.modify_btn = tk.Button(self.form_frame, text="Modify form", command=self.modify_form, fg="darkred", bg="white", font=("Helvetica", 12, "bold"))
            self.modify_btn.grid(row=len(self.fields), column=1, padx=10, pady=20)

            # See the mail
            self.see_mail_btn = tk.Button(self.form_frame, text="See the mail", command=self.see_mail, fg="darkred", bg="white", font=("Helvetica", 12, "bold"))
            self.see_mail_btn.grid(row=len(self.fields) + 1, column=1, padx=10, pady=20)
        
            # Start filling program
            self.start_fill_btn = tk.Button(self.form_frame, text="Start filling program", command=self.fill_documents_2, fg="darkred", bg="white", font=("Helvetica", 12, "bold"))
            self.start_fill_btn.grid(row=len(self.fields) + 2, column=1, padx=10, pady=20)
        

        self.form_canvas.update_idletasks()  # Update the canvas
        self.form_canvas.configure(scrollregion=self.form_canvas.bbox("all"))

    def see_mail(self):
        # Créer une nouvelle fenêtre pour le formulaire de mail
        mail_form = tk.Toplevel(self.root)
        mail_form.title("See mail")
        mail_form.geometry("400x300")

        # Ajouter les champs du formulaire
        tk.Label(mail_form, text="To:").grid(row=0, sticky='e')
        to_entry = tk.Entry(mail_form)
        to_entry.grid(row=0, column=1)
        if 'to' in self.mail_info:
            to_entry.insert(0, self.mail_info['to'])

        tk.Label(mail_form, text="CC:").grid(row=1, sticky='e')
        cc_entry = tk.Entry(mail_form)
        cc_entry.grid(row=1, column=1)
        if 'cc' in self.mail_info:
            cc_entry.insert(0, self.mail_info['cc'])

        tk.Label(mail_form, text="Body:").grid(row=2, sticky='e')
        body_text = tk.Text(mail_form, height=10, width=30)
        body_text.grid(row=2, column=1)
        if 'body' in self.mail_info:
            body_text.insert("1.0", self.mail_info['body'])

    def modify_form(self):
        mdp = simpledialog.askstring("Password", "Enter the password")

        if mdp != self.mdp:
            messagebox.showinfo("Error", "Password incorrect")
            return

        for index, field in enumerate(self.fields):
            field.tkOption.grid(row=index, column=2)
            field.tkDelete.grid(row=index, column=3)
        self.create_form_2("modify")

    def open_settings(self, id):
        # Créer une nouvelle fenêtre ou un cadre pour les paramètres
        self.settings_window = tk.Toplevel(self.root)
        self.settings_window.title("Paramètres")
        self.settings_window.geometry("300x250")  # Définir la taille de la fenêtre

        # Ajouter les boutons dans la fenêtre des paramètres avec le style commun et un espacement vertical
        tk.Button(self.settings_window, text="Is the same valor as", command=lambda: self.link_variable(id), **self.button_style).pack(pady=5)
        tk.Button(self.settings_window, text="Is on the same Excel As", command=lambda: self.set_as_student(id), **self.button_style).pack(pady=5)
        tk.Button(self.settings_window, text="Actual date", command=lambda: self.set_Date_actuelle(id), **self.button_style).pack(pady=5)
        tk.Button(self.settings_window, text="Create your list", command=lambda: self.set_list_of_value(id), **self.button_style).pack(pady=5)
        tk.Button(self.settings_window, text="Remove modification", command=lambda: self.unlink_variable(id), **self.button_style).pack(pady=5)

    def set_as_student(self, id, isExcelObj=None):
        field = self.find_field_with_id(id) 
        
        # Étape 1 : Sélectionner une bookmark "maître"
        if (isExcelObj == None):
            master_id = self.select_variable(id).id
        else:
            master_id = field.isSameExcelAs["masterName"]

        # Étape 2 : Choisir un fichier Excel        
        if (isExcelObj == None):
            excel_path = filedialog.askopenfilename(title="Ouvrir le fichier Excel", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
            if not excel_path:
                return  # L'utilisateur a annulé la sélection du fichier

            # Étape 3 : Demander les noms des colonnes
            self.root.attributes('-topmost', True)
            master_col = simpledialog.askstring("Colonne Maître", "Entrez le chiffre de la colonne 1 = colonne A, 2 = B , ....",parent=self.root)
            self.root.attributes('-topmost', True)
            student_col = simpledialog.askstring("Colonne Élève", "Entrez le chiffre de la colonne 1 = colonne A, 2 = B , ....",parent=self.root)
            self.root.attributes('-topmost', False)
        else:
            excel_path = field.isSameExcelAs["path"]
            master_col = field.isSameExcelAs["masterCol"]
            student_col = field.isSameExcelAs["student_col"]

        if not master_col or not student_col:
            messagebox.showerror("Erreur", "Colonnes non spécifiées.")
            return
        
        master_field = self.find_field_with_id(master_id)

        field.isSameExcelAs = {
            'masterName': master_id,
            'studentName': field.id, 
            'masterCol': int(master_col), 
            'student_col': int(student_col), 
            'path': excel_path
        }
        field.tkInput.config(state=tk.NORMAL)  # Temporairement remettre en état normal pour modifier le texte
        field.tkInput.delete(0, tk.END)
        field.tkInput.insert(0, f"Eleve de {master_field.label} colonne {student_col}")
        field.tkInput.config(state='readonly', readonlybackground='light grey')  # Remettre en état readonly
        self.close_settings()

    def loop(self):
        self.first_page()
        self.root.mainloop()

    def first_page(self):
        # Créer le titre en haut au milieu avec une taille de police plus grande et en gras
        title_label = tk.Label(self.root, text="ConnectAutomate", bg="darkred", fg="white", font=("Helvetica", 24, "bold"), width=20)
        title_label.pack(pady=40)

        # Modifiez le bouton "Create a new form" pour qu'il passe 'root' en argument à 'create_new_form'
        create_new_form_btn = tk.Button(text="Create a new form", command=self.create_new_form, fg="darkred", bg="white", font=("Helvetica", 12, "bold"))
        create_new_form_btn.pack(pady=20, padx=10)

        # Bouton "Load an existing form" avec un contour plus fin
        load_saved_form_btn = tk.Button(text="Load an existing form", command=self.load_saved_form, fg="darkred", bg="white", font=("Helvetica", 12, "bold"))
        load_saved_form_btn.pack(pady=20, padx=10)

    def create_field(self, labelName, path, type, value="", id=-1, isDateToday=False, isSameValue=-1, isSameExcelAs=None, sheetName="", isList=[], pathToSave="", wordNameToSave="", subfolderName="", genericPrefix=-1):
        new_field = Field()

        new_field.label = labelName
        new_field.value = value
        new_field.path = path
        new_field.type = type

        if (id == -1):
            new_field.id = self.currId
            self.currId += 1
        else:
            new_field.id = id
            self.currId = max(self.currId, id) + 1

        new_field.isDateToday = isDateToday
        new_field.isSameValue = isSameValue
        new_field.isSameExcelAs = isSameExcelAs
        new_field.isList = isList

        if sheetName == "":
            new_field.sheetName = self.currSheetName
        else:
            new_field.sheetName = sheetName

        new_field.pathToSave = pathToSave
        new_field.wordNameToSave = wordNameToSave
        new_field.subfolderName = subfolderName
        new_field.genericPrefix = genericPrefix

        self.fields.append(new_field)

    def get_bookmarks_from_docx(self, file_path):
        try:
            document = zipfile.ZipFile(file_path)
            xml_content = document.read('word/document.xml')
            document.close()
            tree = ET.fromstring(xml_content)
            namespace = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
            bookmarks = tree.findall('.//w:bookmarkStart', namespace)
            return [bookmark.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}name') for bookmark in bookmarks]
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la lecture du fichier : {e}")
            return []

    def find_field_with_id(self, id):
        for value in self.fields:
            if value.id == id:
                return value
        return None

    def close_button(self, id):
        field = self.find_field_with_id(id)
        field.tkLabel.destroy()
        field.tkInput.destroy()
        field.tkOption.destroy()
        field.tkDelete.destroy()
        self.fields.remove(field)

    def add_field(self):
        field_obj = self.load_file_new_field()
        self.create_field(field_obj["name"], field_obj["path"], field_obj["type"])
        self.create_form_2("modify")
        
    def set_Date_actuelle(self, id):
        field = self.find_field_with_id(id)

        field.tkInput.config(state=tk.NORMAL)  # Temporairement remettre en état normal pour modifier le texte
        field.tkInput.delete(0, tk.END)
        field.tkInput.insert(0, f"Date du jour")
        field.tkInput.config(state='readonly', readonlybackground='light grey')

        field.isDateToday = True

        self.close_settings()

    def fill_word_with_values(self, references, values, path, pathToSave, wordNameToSave, subfolderName, genericPrefix):
        doc = Document(path)

        for paragraph in doc.paragraphs:
            for index, reference in enumerate(references):
                placeholder = f"{{{{{reference}}}}}"

                if placeholder in paragraph.text:
                    paragraph.text = paragraph.text.replace(placeholder, values[index])


        genericPrefixField = self.find_field_with_id(genericPrefix)
        genericPrefixStr = ""
        if genericPrefixField != None:
            genericPrefixStr = f"_{genericPrefixField.tkInput.get()}"
        
        if subfolderName != "":
            subfolderName = f"{subfolderName}{genericPrefixStr}"
    

        file_path = os.path.join(f"{pathToSave}/{subfolderName}", f"{wordNameToSave}{genericPrefixStr}.docx")
        if not os.path.exists(f"{pathToSave}/{subfolderName}") and subfolderName != "":
            os.makedirs(f"{pathToSave}/{subfolderName}")

        print(pathToSave, subfolderName, genericPrefixStr, wordNameToSave, file_path)

        doc.save(file_path)
            
    def fill_excel_with_values(self, references, values, path, sheetName):
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        add_row = []

        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            if cell_value in references:  # Vérifier si cell_value est dans references
                index = references.index(cell_value)  # Obtenir l'index seulement si cell_value est trouvé
                add_row.append(values[index])  # Utiliser l'index pour ajouter la valeur correspondante à add_row

        
        ws.append(add_row)
        wb.save(path)

    def get_all_paths_with_type(self):
        paths = []
        for field in self.fields:
            path_type = {"path": field.path, "type": field.type, "pathToSave": field.pathToSave, "wordNameToSave": field.wordNameToSave, "subfolderName": field.subfolderName, "genericPrefix": field.genericPrefix}
            if path_type not in paths:
                paths.append(path_type)
        return paths

    def fill_documents_2(self):
        paths_types = self.get_all_paths_with_type()

        for path_type in paths_types:
            path = path_type["path"]

            references = []
            values = []
            for field in self.fields:
                if field.path == path:
                    field_use = field

                    reference = field_use.label

                    if field_use.isSameValue != -1:
                        field_link = self.find_field_with_id(field_use.isSameValue)
                        while field_link.isSameValue != -1:
                            field_link = self.find_field_with_id(field_link.isSameValue)
                        field_use = field_link

                    value = field_use.tkInput.get()

                    if field_use.isDateToday:
                        value = datetime.datetime.now().strftime("%d/%m/%Y")
                    
                    if field_use.isSameExcelAs != None:
                        master_field = self.find_field_with_id(field_use.isSameExcelAs["masterName"])
                        student_field = self.find_field_with_id(field_use.isSameExcelAs["studentName"])
                        master_col = field_use.isSameExcelAs["masterCol"] - 1
                        student_col = field_use.isSameExcelAs["student_col"] - 1

                        wb = openpyxl.load_workbook(field_use.isSameExcelAs["path"])
                        ws = wb.active

                        for row in ws.iter_rows(values_only=True):
                            a = row[master_col] 
                            b = master_field.tkInput.get()
                            if str(row[master_col]) == str(master_field.tkInput.get()):
                                reference = student_field.label
                                value = row[student_col]

                        wb.close()

                    references.append(reference)
                    values.append(value)

            print("REFERENCES: ", references)
            print("VALUES    : ", values)

            if path_type["type"] == "Word":
                self.fill_word_with_values(references, values, path, path_type['pathToSave'], path_type['wordNameToSave'], path_type['subfolderName'], path_type['genericPrefix'])

            if path_type["type"] == "Excel":
                self.fill_excel_with_values(references, values, path, "")

        messagebox.showinfo("Succès", "Les document ont été remplis et sauvegardés.")

    def link_variable(self, id, id_field_link=-1):
        field_actual = self.find_field_with_id(id)
        if (id_field_link == -1):
            field_link = self.select_variable(id)
            field_actual.isSameValue = field_link.id
        else:
            field_actual.isSameValue = id_field_link

        field_link = self.find_field_with_id(field_actual.isSameValue)

        field_actual.tkInput.config(state=tk.NORMAL)
        field_actual.tkInput.delete(0, tk.END)
        field_actual.tkInput.insert(0, f"Lié à {field_link.label}")
        field_actual.tkInput.config(state='readonly', readonlybackground='light grey')

        self.close_settings()

    def unlink_variable(self, id):
        field = self.find_field_with_id(id)

        field.tkInput.config(state='normal', relief=tk.SUNKEN)
        field.tkInput.delete(0, tk.END)
        self.close_settings()
    
    def set_list_of_value(self, id, idList=[]):
        # Créer une nouvelle fenêtre pour la liste
        field = self.find_field_with_id(id)

        # Création de la combobox dans le form_frame
        if hasattr(field, 'tkInput'):
            field.tkInput.grid_forget()  # Cache le tkInput sans le détruire

        combobox_main = ttk.Combobox(self.form_frame, state="readonly")
        combobox_main.grid(row=field.id, column=1)  # Placer la combobox au même endroit que tkInput
        field.tkInput = combobox_main  # Remplacer tkInput par la combobox

        # Initialiser la combobox avec un élément vide ou un texte de prompt
        combobox_main['values'] = tuple(idList) if idList else ("Select an item",)  # Initialiser avec les valeurs fournies ou un texte de prompt
        combobox_main.set(idList[0] if idList else "Select an item")

        # Si idList est vide, permettre à l'utilisateur d'ajouter des valeurs à la liste
        if not idList:
            list_window = tk.Toplevel(self.settings_window)
            list_window.title("Create your list")
            list_window.geometry("350x150")

            # Créer une combobox non éditable
            combobox_values = tk.StringVar()
            combobox_list_window = ttk.Combobox(list_window, textvariable=combobox_values, state="readonly")
            combobox_list_window.pack(pady=(10, 20))  # Ajouter un espace après la combobox

            # Zone de texte et bouton pour ajouter des éléments directement dans list_window
            new_field_var = tk.StringVar(value="Write your next field")
            new_field_entry = tk.Entry(list_window, textvariable=new_field_var)
            new_field_entry.pack(pady=(10, 0))

            def submit_new_field():
                new_field_value = new_field_var.get()
                if new_field_value and new_field_value not in combobox_list_window['values']:
                    combobox_list_window['values'] = (*combobox_list_window['values'], new_field_value)
                    new_field_var.set("Write your next field")  # Réinitialiser la zone de texte

            submit_button = tk.Button(list_window, text="Add to List", command=submit_new_field, **self.button_style)
            submit_button.pack(pady=(10, 0))

            # Bouton pour enregistrer la liste et fermer les fenêtres
            def save_list_and_close():
                # Mettre à jour field.isList avec les valeurs de combobox_list_window
                field.isList = list(combobox_list_window['values'])
                combobox_main['values'] = field.isList  # Mettre à jour la combobox principale avec les nouvelles valeurs
                if field.isList:
                    combobox_main.set(field.isList[0])  # Sélectionnez le premier élément si la liste n'est pas vide
                else:
                    combobox_main.set("Select an item")  # Ou laissez vide si la liste est vide
                list_window.destroy()

            tk.Button(list_window, text="Save your list", command=save_list_and_close, **self.button_style).pack()

    def confirm_selection(self, field, win):
        self.tmp_value = field
        win.destroy()

    def select_variable(self, id):
        selection_win = tk.Toplevel(self.root)
        selection_win.title("Sélectionner la variable à lier")

        for index, field in enumerate(self.fields):
            if field.id != id:
                btn = tk.Button(selection_win, text=field.label, command=lambda field=field: self.confirm_selection(field, selection_win))
                btn.grid(row=index, column=0)

        selection_win.grab_set()
        self.root.wait_window(selection_win)

        return self.tmp_value

    def get_xlsx_column_name_filtered(self, file_path):
        first_row_values = self.get_xlsx_column_name(file_path)
        self.root.attributes('-topmost', True)
        first_column = int(simpledialog.askstring("Sélection de colonnes excel", "Selectionner colonne inital, (1=A, 2=B, ...)",parent = self.root))
        self.root.attributes('-topmost', True)
        last_column  = int(simpledialog.askstring("Sélection de colonnes excel", "Selectionner colonne final, (1=A, 2=B, ...)",parent = self.root))
        self.root.attributes('-topmost', False)

        values_filtered = []
        for i in range(first_column, last_column + 1):
            if (len(first_row_values) >= i):
                values_filtered.append(first_row_values[i - 1])

        return values_filtered

    def get_xlsx_column_name(self, file_path):
        workbook = openpyxl.load_workbook(file_path)
        if len(workbook.sheetnames) == 1:
            sheet = workbook.active
        else:
            dialog = tk.Toplevel(self.root)
            dialog.title("Select sheet")

            def get_selected_value():
                self.tmp_value = combobox.get()
                dialog.destroy()

            selected_value = tk.StringVar()
            combobox = Combobox(dialog, values=workbook.sheetnames, textvariable=selected_value)
            combobox.pack(padx=20, pady=10)

            close_button = tk.Button(dialog, text="OK", command=get_selected_value)
            close_button.pack(pady=10)

            dialog.wait_window()
            sheet = workbook[self.tmp_value]
            self.currSheetName = self.tmp_value
                
        first_row_values = []
        for cell in sheet[1]:  # Iterate over cells in the first row
            first_row_values.append(cell.value)

        return first_row_values

    def load_file_new_field(self):
        file_path = filedialog.askopenfilename(filetypes=[
            ("All files", "*.*"),
            ("Word files", "*.docx"),
            ("Excel files", "*.xlsx")
        ])

        file_extension = os.path.splitext(file_path)[1].lower()

        field_obj = {"name": "", "type": "", "path": file_path}

        if file_path and file_extension == ".xlsx":
            field_obj["type"] = "Excel"
            column_names = self.get_xlsx_column_name(file_path)
            column_id = int(simpledialog.askstring("Sélection une colonne excel", "Selectionner la colonne, (1=A, 2=B, ...)"))
            
            if (column_id - 1 < len(column_names)):
                field_obj["name"] = column_names[column_id - 1]


        if file_path and file_extension == ".docx":
            field_obj["type"] = "Word"
            bookmark_names = self.get_bookmarks_from_docx(file_path)

            dialog = tk.Toplevel(self.root)
            dialog.title("Select signet")

            def get_selected_value():
                self.tmp_value = combobox.get()
                dialog.destroy()

            selected_value = tk.StringVar()
            combobox = Combobox(dialog, values=bookmark_names, textvariable=selected_value)
            combobox.pack(padx=20, pady=10)

            close_button = tk.Button(dialog, text="OK", command=get_selected_value)
            close_button.pack(pady=10)

            dialog.wait_window()
            field_obj["name"] = self.tmp_value
        
        return field_obj

    def load_file(self):
        file_path = filedialog.askopenfilename(filetypes=[
            ("All files", "*.*"),
            ("Word files", "*.docx"),
            ("Excel files", "*.xlsx")
        ])

        file_extension = os.path.splitext(file_path)[1].lower()

        if file_path and file_extension == ".xlsx":
            excel_column_names = self.get_xlsx_column_name_filtered(file_path)
            if len(excel_column_names) == 0:
                messagebox.showinfo("Information", "Aucun informations existant dans ce fichier.")

            for column_name in excel_column_names:
                self.create_field(column_name, file_path, "Excel")
            self.create_form_2("modify")


        if file_path and file_extension == ".docx":
            bookmark_names = self.get_bookmarks_from_docx(file_path)
            if len(bookmark_names) == 0:
                messagebox.showinfo("Information", "Aucun signet existant dans ce fichier.")

            for bookmark_name in bookmark_names:
                self.create_field(bookmark_name, file_path, "Word")
            self.create_form_2("modify")

    def reset_to_menu(self):
        for widget in self.root.winfo_children():
            widget.destroy()
        self.first_page()

        self.fields = []
        self.askToSave = False

        self.load_word_file_btn = None
        self.add_field_btn = None
        self.add_mail_btn = None
        self.save_btn = None
        self.modify_btn = None
        self.see_mail_btn = None
        self.start_fill_btn = None
        self.form_frame = None

    def create_save_menu(self):
        user_choice = messagebox.askyesno("Save", "Are you sure to leave without save?")
        return user_choice

    def go_back(self):
        if (self.askToSave):
            if (self.create_save_menu()):
                self.reset_to_menu()
        else:
            self.reset_to_menu()

    def close_settings(self):
        # Fermer la fenêtre des paramètres
        if self.settings_window != {}:
            self.settings_window.destroy()

    def add_mail(self):
        # Vérifier si un mail est déjà ajouté et initialiser le formulaire en conséquence
        if hasattr(self, 'mail_info') and self.mail_info:
            self.add_mail_btn.config(text="Modify mail")
        else:
            self.mail_info = {}

        # Créer une nouvelle fenêtre pour le formulaire de mail
        mail_form = tk.Toplevel(self.root)
        mail_form.title("Add/Modify Mail")
        mail_form.geometry("400x300")

        # Ajouter les champs du formulaire
        tk.Label(mail_form, text="To:").grid(row=0, sticky='e')
        to_entry = tk.Entry(mail_form)
        to_entry.grid(row=0, column=1)
        if 'to' in self.mail_info:
            to_entry.insert(0, self.mail_info['to'])

        tk.Label(mail_form, text="CC:").grid(row=1, sticky='e')
        cc_entry = tk.Entry(mail_form)
        cc_entry.grid(row=1, column=1)
        if 'cc' in self.mail_info:
            cc_entry.insert(0, self.mail_info['cc'])

        tk.Label(mail_form, text="Body:").grid(row=2, sticky='e')
        body_text = tk.Text(mail_form, height=10, width=30)
        body_text.grid(row=2, column=1)
        if 'body' in self.mail_info:
            body_text.insert("1.0", self.mail_info['body'])

        # Bouton pour sauvegarder le mail
        save_mail_btn = tk.Button(mail_form, text="Save Mail", command=lambda: self.save_mail(to_entry, cc_entry, body_text))
        save_mail_btn.grid(row=3, column=1, pady=10)

        # Bouton pour supprimer le mail
        remove_mail_btn = tk.Button(mail_form, text="Remove Mail", command=self.remove_mail)
        remove_mail_btn.grid(row=4, column=1, pady=10, sticky="ew")

    def save_mail(self, to_entry, cc_entry, body_text):
        # Sauvegarder les informations du mail
        self.mail_info['to'] = to_entry.get()
        self.mail_info['cc'] = cc_entry.get()
        self.mail_info['body'] = body_text.get("1.0", tk.END)

        # Changer le texte du bouton pour refléter que le mail a été ajouté/modifié
        self.add_mail_btn.config(text="Modify mail")

        self.askToSave = True

        # Fermer le formulaire de mail
        self.root.focus_force()

    def send_mail(to, cc, body):
        from_addr = "your_email@example.com"
        msg = MIMEMultipart()
        msg['From'] = from_addr
        msg['To'] = to
        msg['Cc'] = cc
        msg

    def remove_mail(self):
        # Vider les informations du mail
        self.mail_info.clear()

        # Changer le texte du bouton pour indiquer que le mail peut être ajouté
        self.add_mail_btn.config(text="Add a mail")

        self.askToSave = False

        # Fermer le formulaire de mail et revenir à la fenêtre principale
        self.root.focus_force()

    def create_new_form(self):
        for widget in self.root.winfo_children():
            widget.destroy()

        # Créer un conteneur pour les boutons
        buttons_frame = tk.Frame()
        buttons_frame.pack(fill='x', pady=20)

        # Bouton retour avec une flèche (ou texte)
        back_btn = tk.Button(buttons_frame, text="←", command=self.go_back)
        back_btn.grid(row=0, column=0, sticky='w', padx=0)

        # Espaceur pour centrer le deuxième bouton
        buttons_frame.grid_columnconfigure(1, weight=1)

        # Lors de la création des boutons, assignez-les comme attributs de l'instance de la classe
        self.load_word_file_btn = tk.Button(buttons_frame, text="Automatic filling from a (Word, Excel) file", command=self.load_file, fg="darkred", bg="white", font=("Helvetica", 12, "bold"))
        self.load_word_file_btn.grid(row=20, column=1, padx=10, pady=20)

        self.add_field_btn = tk.Button(buttons_frame, text="+ Add a new field", command=self.add_field, fg="darkred", bg="white", font=("Helvetica", 12, "bold"))
        self.add_field_btn.grid(row=21, column=1, padx=10)

        self.add_mail_btn = tk.Button(buttons_frame, text="Add a mail", command=self.add_mail, fg="darkred", bg="white", font=("Helvetica", 12, "bold"))
        self.add_mail_btn.grid(row=22, column=1, padx=10, pady=20)


form = BookmarkForm()
form.loop()

print("END")